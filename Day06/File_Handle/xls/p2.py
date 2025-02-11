# TO LOAD THE CONTENT FROM XLS

from openpyxl import load_workbook
wb = load_workbook("p1.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)